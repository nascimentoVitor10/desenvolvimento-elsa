import pandas as pd
import polars as pl
import sys
from openpyxl import load_workbook


# pipeline de extração de dados

# [1]
def save_vars_in_txt(vars: list, txt_file_path: str):
    """
    Essa função irá acessar o caminho do arquivo passado
    e irá preenchê-lo com as variáveis informadas para que
    elas possam ser salvas para posteriores consultas

    Parâmetros:

        vars(list)      : variáveis que ainda não foram encontradas nos dados
        file_path(str)  : caminho do arquivo .txt em que as serão salvas as variáveis
    """
    try:
        with open(txt_file_path, 'w') as file:
            for var in vars:
                file.write(var + '\n')
            print("arquivo preenchido!")

    except IOError:
        print("arquivo não encontrado")


# [2]
def load_list_txt(txt_file_path: str) -> list:
    """
    Retorna uma lista python das variáveis salvas no arquivo .txt

    Parâmetros:
        file_path: caminho do arquivo em que as variáveis estão armazenadas
    """
    vars = []
    with open(txt_file_path, 'r') as file:
        for line in file:
            if "\n" in line:
                line = line.replace('\n', '')
                vars.append((line.strip()).upper())
    return vars


# [3]
def get_columns_var_excel_file(file_path: str) -> list:
    """
    Retorna uma lista de variáveis que compõem as colunas de um arquivo .xslx,
    pois os arquivos excel são mais trabalhosos para serem carregados quando
    muito grandes.

    Parâmetros:
        file_path: caminho do arquivo .xslx a ser analisado

    """
    vars = []
    workbook = load_workbook(filename=file_path, read_only=True)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
        vars = [i.upper() for i in row]

    return vars


# [4]
def compare_lists(df_vars: list, txt_vars: list) -> tuple:
    """
    Retorna uma lista de variáveis que estão presentes no dataframe e
    atualiza o arquivo .txt em que as variáveis não encontradas no
    dataframe ficam armazenadas.

    Parâmetros:
        df_vars(list) : lista de variáveis que compõem as colunas do dataframe
        txt_vars(list): lista das variáveis presentes no arquivo .txt
    """
    txt_file_path = r'E:\elsa\ELSA\desenvolvimento-elsa\scripts\elsa_organizacao_bases\dados_extracao\vars_n_encontradas.txt'
    if 'IDELSA' not in txt_vars:
        txt_vars.insert(0, 'IDELSA')

    variaveis_encontradas_df = [i for i in txt_vars if i in df_vars]
    variaveis_Nencontradas_df = [i for i in txt_vars if i not in df_vars]

    save_vars_in_txt(variaveis_Nencontradas_df, txt_file_path)

    return variaveis_encontradas_df


def save_df(df: pd.DataFrame, id: str, path_file: str):
    df.to_excel(path_file)
    print(f'df_{id} : salvo')
    print(f'    rows:{df.shape[0]} -- cols:{df.shape[1]}\n')
    del df


# [5]
def load_large_dta(file_path: str, txt_file_path: str) -> pd.DataFrame:
    """
    Retorna dataframe correspondente ao caminho de arquivo .dta informado

    Parâmetros:
        file_path: caminho do arquivo .dta a ser analisado
    """
    reader = pd.read_stata(file_path, iterator=True)
    df = pd.DataFrame()
    try:
        chunk = reader.get_chunk(100 * 1000)
        while len(chunk) > 0:
            df = pd.concat([df, chunk], ignore_index=True)
            chunk = reader.get_chunk(100 * 1000)
            sys.stdout.flush()
    except (StopIteration, KeyboardInterrupt):
        pass

    df_columns_upper = [i.upper() for i in df.columns]
    df.columns = df_columns_upper
    print(f'df : {df.columns}')

    vars_in_txt = load_list_txt(txt_file_path=txt_file_path)
    vars_in_df = compare_lists(list(df.columns), vars_in_txt)

    if len(vars_in_df) <= 1:
        print("Nenhuma variável encontrada nessa base.")
        return pd.DataFrame(None)

    return df[vars_in_df].sort_values(by='IDELSA', ascending=True)


# [6]
def load_large_excel(file_path: str, txt_file_path: str) -> pd.DataFrame:
    """
    Retorna o dataframe do arquivo .xslx já filtrado com as variáveis encontradas

    Parâmetros:

        file_path    : caminho do arquivo .xslx a ser analisado
        txt_file_path: caminho do arquivo .txt em que as variáveis ficam armazenadas
    """
    vars_df = get_columns_var_excel_file(file_path=file_path)
    vars_file_txt = load_list_txt(txt_file_path=txt_file_path)
    vars_in_df = compare_lists(vars_df, vars_file_txt)

    if len(vars_in_df) <= 1:
        print("Nenhuma variável encontrada nessa base.")
        return pd.DataFrame(None)

    df_filtrado = pd.read_excel(file_path, engine='calamine', usecols=vars_in_df)
    return df_filtrado.sort_values(by='IDELSA', ascending=True)


def main():

    lista_dirs1={
        'base01_completa'    : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - COMPLETA\O1.xlsx",
        'laboratorio'        : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Laboratório Onda 1\base_lab_onda1_151120.dta" ,
        'labSI'              : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Laboratório Onda 1\Base Laboratório Sistema Internacional Onda 1\base_LAB_SI_onda1_151120.dta",
        'medContinua_pa'     : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Medicamentos Onda 1\Continuo\base_med_o1_c_pa_150619.dta",
        'medContinua_p1'     : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Medicamentos Onda 1\Continuo\base_med_o1_c_150619_parte1.dta",
        'medContinua_p2'     : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Medicamentos Onda 1\Continuo\base_med_o1_c_150619_parte2.dta",
        'medContinua_p3'     : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Medicamentos Onda 1\Continuo\base_med_o1_c_150619_parte3.dta",
        'dieta'              : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Dieta Onda 1\base_dieta_onda1_150619.dta",
        'der_sindr_met'      : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1_Base Derivadas_Síndrome Metabólica_JUL2015\base_derivadas_090715.xlsx",
        'hemograma'          : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Hemograma Onda1\Base_Hemograma_onda1_150619.dta",
        'georeferenciamento' : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Geo Onda 1 Nova\Base_Geo_onda1_150619.dta",
        'ocupação'           : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Ocupação Onda 1\base_ocupacao_150619_idelsa.dta",
        'ultraprocessadps'   : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Ultraprocessados onda 1\base_var_ultraprocessados_140319.dta"
        }

    # lista_dirs2={
        # 'base02_completa': r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 2\O2 - Base Completa Onda 2 Nova\base_elsa_onda2_150619.dta",
        # 'bioimpd'        : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 2\O2 - Base Bioimpedancia Onda 2\base_bio_onda2_150619_dist.dta",
        # 'lab'            : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 2\O1 - Base Laboratório Onda 2\base_lab_onda2_151120.dta",
        # 'labSI'          : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 2\O1 - Base Laboratório Onda 2\Base Laboratório Sistema Internacional Onda 2\base_LAB_SI_onda2_151120.dta",
        # 'medClasses'     : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 2\O2 - Base Medicamentos Onda 2\Classes\BASE_MED_O2_150619_DIST.dta",
        # 'medConcent_A_C' : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 2\O2 - Base Medicamentos Onda 2\Concentracao\MED_O2_CON_A_C_DIST_150619.dta",
        # 'medConcent_D_K' : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 2\O2 - Base Medicamentos Onda 2\Concentracao\MED_O2_CON_D_K_DIST_150619.dta",
        # 'medConcent_L_P' : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 2\O2 - Base Medicamentos Onda 2\Concentracao\MED_O2_CON_L_P_DIST_150619.dta",
        # 'medConcent_Q_Z' : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 2\O2 - Base Medicamentos Onda 2\Concentracao\MED_O2_CON_Q_Z_DIST_150619.dta" ,
        # 'hemo'           : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Dieta Onda 1\base_dieta_onda1_150619.dta",
        # 'PCR'            : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1_Base Derivadas_Síndrome Metabólica_JUL2015\base_derivadas_090715.xlsx",
        # 'RETCL'          : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Hemograma Onda1\Base_Hemograma_onda1_150619.dta",
        # 'recusa_obito'   : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Geo Onda 1 Nova\Base_Geo_onda1_150619.dta",
        # 'derivadas'      : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Ultraprocessados onda 1\base_var_ultraprocessados_140319.dta"
        #}

        # lista_dirs3={
        # 'base01_completa': r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - COMPLETA\O1.xlsx",
        # 'lab'            : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Laboratório Onda 1\base_lab_onda1_151120.dta" ,
        # 'labSI'          : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Laboratório Onda 1\Base Laboratório Sistema Internacional Onda 1\base_LAB_SI_onda1_151120.dta",
        # 'medCont_p1'     : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Medicamentos Onda 1\Continuo\base_med_o1_c_150619_parte1.dta",
        # 'medCont_p2'     : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Medicamentos Onda 1\Continuo\base_med_o1_c_150619_parte2.dta",
        # 'medCont_p3'     : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Medicamentos Onda 1\Continuo\base_med_o1_c_150619_parte3.dta",
        # 'dieta'          : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Dieta Onda 1\base_dieta_onda1_150619.dta",
        # 'der_dieta'      : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1_Base Derivadas_Síndrome Metabólica_JUL2015\base_derivadas_090715.xlsx",
        # 'hemo'           : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Hemograma Onda1\Base_Hemograma_onda1_150619.dta",
        # 'geo'            : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Geo Onda 1 Nova\Base_Geo_onda1_150619.dta",
        # 'ocup'           : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Base Ocupação Onda 1\base_ocupacao_150619_idelsa.dta",
        # 'ultrap'         : r"E:\elsa\ELSA\Bases de dados\Base Elsa\Bases - onda 1\O1 - Ultraprocessados onda 1\base_var_ultraprocessados_140319.dta"
        #}


    path_dict = r"E:\elsa\ELSA\Solicitações\Solicitações Vitor Ramos\Oscar Martinez\08_11_24\Dicionario_ELSA_Milena_Santos.xlsx"
    df_pedido = pd.read_excel(path_dict)
    df_pedido_onda1 = df_pedido.iloc[:, 0:3]
    df_pedido_onda2 = df_pedido.iloc[:, 4:7]
    df_pedido_onda3 = df_pedido.iloc[:, 8:11]

    variaveis_pedidas_onda1 = list(set([i.strip().upper() for i in df_pedido_onda1[df_pedido_onda1['Unnamed: 1'].notnull()].iloc[1:, 1]]))
    variaveis_pedidas_onda2 = list(set([i.strip().upper() for i in df_pedido_onda2[df_pedido_onda2['Unnamed: 5'].notnull()].iloc[1:, 1]]))
    variaveis_pedidas_onda3 = list(set([i.strip().upper() for i in df_pedido_onda3[df_pedido_onda3['Unnamed: 9'].notnull()].iloc[1:, 1]]))


    txt_file_path = r'E:\elsa\ELSA\desenvolvimento-elsa\scripts\elsa_organizacao_bases\extracao\vars_encontrar.txt'
    save_vars_in_txt(vars=variaveis_pedidas_onda1, txt_file_path=txt_file_path)
    with pd.ExcelWriter(r'E:\elsa\ELSA\desenvolvimento-elsa\scripts\elsa_organizacao_bases\extracao\onda2\extracao_onda1.xlsx', engine='xlsxwriter') as writer:

        for idx, base in enumerate(lista_dirs1):

            if lista_dirs1[base].endswith('.dta'):
                df = load_large_dta(file_path=lista_dirs1[base], txt_file_path=txt_file_path)
        
            else:
                df = load_large_excel(file_path=lista_dirs1[base], txt_file_path=txt_file_path)
        
            df.to_excel(writer, sheet_name=base + f"- tabela {idx + 1}")
            del df

if __name__ == '__main__':
    main()



